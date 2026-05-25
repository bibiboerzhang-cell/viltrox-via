#!/usr/bin/env python3
"""Prepare Viltrox promotion-plan XLSX data for KOL Pool import.

Default mode is read-only: parse the sheet, export cleaned owner/KOL mapping,
and write a JSON payload that can be reviewed before any DB import.
"""
from __future__ import annotations

import argparse
import csv
import json
import re
import sys
from collections import Counter
from datetime import datetime, UTC
from pathlib import Path
from typing import Any
from urllib.parse import urlparse

import openpyxl


OWNER_COL = "登记/对接人"
KOL_COL = "红人/媒体"
PLATFORM_COL = "平台"
PRODUCT_COL = "推广产品"
VIDEO_URL_COL = "红人视频链接"
DELIVERED_URL_COL = "回片链接"
FOLLOWERS_COL = "粉丝数/访客数"
COUNTRY_COL = "国家"
STATUS_COL = "合作进度"
ENGAGEMENT_COL = "互动率"


def _text(value: Any) -> str:
    return str(value or "").strip()


def _slug(value: str) -> str:
    return re.sub(r"[^a-z0-9_@.-]+", "_", value.strip().lower()).strip("_")


def _platform(value: str) -> str:
    raw = value.strip().lower().replace("，", ",")
    if not raw:
        return "other"
    if "instagram" in raw or raw == "ig":
        return "instagram"
    if "tiktok" in raw or "tik tok" in raw:
        return "tiktok"
    if "youtube" in raw or raw in {"yt", "youtu"}:
        return "youtube"
    if "media" in raw:
        return "media"
    if "facebook" in raw:
        return "facebook"
    if raw in {"x", "twitter"}:
        return "x"
    return raw.split(",")[0].strip() or "other"


def _parse_number(value: Any) -> int | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return int(value)
    raw = str(value).strip().replace(",", "")
    if not raw:
        return None
    multiplier = 1
    lowered = raw.lower()
    if lowered.endswith("k"):
        multiplier = 1_000
        raw = raw[:-1]
    elif lowered.endswith("m"):
        multiplier = 1_000_000
        raw = raw[:-1]
    elif lowered.endswith("w") or raw.endswith("万"):
        multiplier = 10_000
        raw = raw[:-1]
    try:
        return int(float(raw) * multiplier)
    except ValueError:
        return None


def _parse_rate(value: Any) -> float | None:
    raw = _text(value)
    if not raw:
        return None
    try:
        if raw.endswith("%"):
            return float(raw[:-1]) / 100
        return float(raw)
    except ValueError:
        return None


def _extract_handle_from_url(url: str, platform: str) -> str:
    if not url:
        return ""
    parsed = urlparse(url if re.match(r"^https?://", url) else f"https://{url}")
    path_parts = [part for part in parsed.path.split("/") if part]
    if not path_parts:
        return ""
    if platform == "instagram":
        return path_parts[0].lstrip("@")
    if platform == "tiktok":
        for part in path_parts:
            if part.startswith("@"):
                return part.lstrip("@")
        return path_parts[0].lstrip("@")
    if platform == "youtube":
        for index, part in enumerate(path_parts):
            if part.startswith("@"):
                return part.lstrip("@")
            if part in {"channel", "c", "user"} and index + 1 < len(path_parts):
                return path_parts[index + 1]
    return path_parts[0].lstrip("@")


def _guess_handle(row: dict[str, Any], platform: str) -> str:
    for column in (VIDEO_URL_COL, DELIVERED_URL_COL):
        handle = _extract_handle_from_url(_text(row.get(column)), platform)
        if handle:
            return handle
    return _slug(_text(row.get(KOL_COL))).lstrip("@")


def _profile_url(platform: str, handle: str, fallback_url: str) -> str:
    if not handle:
        return fallback_url
    if platform == "instagram":
        return f"https://www.instagram.com/{handle}/"
    if platform == "tiktok":
        return f"https://www.tiktok.com/@{handle}"
    if platform == "youtube":
        if handle.startswith("UC"):
            return f"https://www.youtube.com/channel/{handle}"
        return f"https://www.youtube.com/@{handle}"
    return fallback_url


def load_rows(path: Path) -> list[dict[str, Any]]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    headers = [_text(cell.value) for cell in ws[1]]
    rows: list[dict[str, Any]] = []
    for values in ws.iter_rows(min_row=2, values_only=True):
        row = {headers[index]: value for index, value in enumerate(values) if index < len(headers)}
        if any(_text(value) for value in row.values()):
            rows.append(row)
    return rows


def build_items(rows: list[dict[str, Any]], *, source_ref: str) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]], dict[str, Any]]:
    entities: dict[tuple[str, str], dict[str, Any]] = {}
    mapping_rows: list[dict[str, Any]] = []
    collaboration_records: list[dict[str, Any]] = []
    owner_counter: Counter[str] = Counter()
    platform_counter: Counter[str] = Counter()
    product_counter: Counter[str] = Counter()
    status_counter: Counter[str] = Counter()
    seen: set[tuple[str, str]] = set()
    duplicates = 0

    for index, row in enumerate(rows, start=2):
        owner = _text(row.get(OWNER_COL)) or "未分配"
        display_name = _text(row.get(KOL_COL))
        platform = _platform(_text(row.get(PLATFORM_COL)))
        video_url = _text(row.get(VIDEO_URL_COL))
        delivered_url = _text(row.get(DELIVERED_URL_COL))
        handle = _guess_handle(row, platform)
        if not display_name and not handle:
            continue
        key = (platform, handle or _slug(display_name))
        if key in seen:
            duplicates += 1
        seen.add(key)

        product = _text(row.get(PRODUCT_COL))
        status = _text(row.get(STATUS_COL))
        followers = _parse_number(row.get(FOLLOWERS_COL))
        engagement_rate = _parse_rate(row.get(ENGAGEMENT_COL))
        profile_url = _profile_url(platform, handle, video_url or delivered_url)
        record = {
            "xlsx_row": index,
            "source_scope": "partial",
            "owner_name": owner,
            "kol_name": display_name or handle,
            "platform": platform,
            "handle": handle,
            "product": product,
            "status": status,
            "country": _text(row.get(COUNTRY_COL)),
            "profile_url": profile_url,
            "video_url": video_url,
            "delivered_url": delivered_url,
            "followers": followers,
            "engagement_rate": engagement_rate,
            "source_columns": {key: _text(value) for key, value in row.items()},
        }

        owner_counter[owner] += 1
        platform_counter[platform] += 1
        if product:
            product_counter[product] += 1
        if status:
            status_counter[status] += 1

        entity = entities.get(key)
        if not entity:
            entity = {
                "platform": platform,
                "handle": handle,
                "display_name": display_name or handle,
                "profile_url": profile_url,
                "followers": followers,
                "engagement_rate": engagement_rate,
                "source_type": "promo_plan_xlsx",
                "source_ref": source_ref,
                "source_scope": "partial",
                "owner_names": [],
                "country": _text(row.get(COUNTRY_COL)),
                "recommended_product_lines": [],
                "raw": {
                    "source_scope": "partial",
                    "source_file_is_complete_roster": False,
                    "source_file": source_ref,
                    "entity_key": f"{platform}:{handle or _slug(display_name)}",
                    "collaboration_records": [],
                },
            }
            entities[key] = entity
        if owner and owner not in entity["owner_names"]:
            entity["owner_names"].append(owner)
        if product and product not in entity["recommended_product_lines"]:
            entity["recommended_product_lines"].append(product)
        if followers is not None:
            entity["followers"] = followers
        if engagement_rate is not None:
            entity["engagement_rate"] = engagement_rate
        if profile_url and not entity.get("profile_url"):
            entity["profile_url"] = profile_url
        entity["raw"]["collaboration_records"].append(record)
        collaboration_records.append(record)
        mapping_rows.append(
            {
                "xlsx_row": index,
                "owner": owner,
                "kol_name": display_name or handle,
                "platform": platform,
                "handle": handle,
                "product": product,
                "status": status,
                "profile_url": profile_url,
                "video_url": video_url,
                "delivered_url": delivered_url,
                "followers": followers or "",
                "engagement_rate": engagement_rate if engagement_rate is not None else "",
                "source_scope": "partial",
            }
        )

    items = list(entities.values())
    summary = {
        "source_rows": len(rows),
        "kol_entity_items": len(items),
        "collaboration_records": len(collaboration_records),
        "unique_platform_handle": len(seen),
        "duplicate_platform_handle_rows": duplicates,
        "source_scope": "partial",
        "source_file_is_complete_roster": False,
        "owners": dict(owner_counter.most_common()),
        "platforms": dict(platform_counter.most_common()),
        "products": dict(product_counter.most_common()),
        "statuses": dict(status_counter.most_common()),
    }
    return items, mapping_rows, collaboration_records, summary


def write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if not rows:
        path.write_text("", encoding="utf-8")
        return
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(rows[0].keys()))
        writer.writeheader()
        writer.writerows(rows)


def apply_kol_pool(items: list[dict[str, Any]], staff_id: int, source_ref: str) -> dict[str, Any]:
    sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "backend"))
    from app.domains.kol import pool as kol_pool

    payload_items = []
    for item in items:
        raw = dict(item)
        raw.update(item.get("raw") or {})
        raw["source_scope"] = "partial"
        raw["source_file_is_complete_roster"] = False
        payload_items.append(raw)
    return kol_pool.import_items(
        payload_items,
        source_type="promo_plan_xlsx",
        source_ref=source_ref,
        staff={"id": staff_id, "role": "admin", "is_owner": 1},
    )


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--xlsx", required=True, help="Promotion plan xlsx path")
    parser.add_argument("--out-dir", default="", help="Output directory")
    parser.add_argument("--apply-kol-pool", action="store_true", help="Import cleaned rows into vkpi_kol_pool")
    parser.add_argument("--staff-id", type=int, default=0, help="Audit staff id for --apply-kol-pool")
    args = parser.parse_args()

    xlsx_path = Path(args.xlsx).expanduser()
    if not xlsx_path.exists():
        raise SystemExit(f"xlsx not found: {xlsx_path}")

    stamp = datetime.now(UTC).strftime("%Y%m%d-%H%M%S")
    out_dir = Path(args.out_dir or f"runtime/qa/promo-plan-import-{stamp}").resolve()
    rows = load_rows(xlsx_path)
    items, mapping_rows, collaboration_records, summary = build_items(rows, source_ref=xlsx_path.name)
    summary["source_file"] = str(xlsx_path)
    summary["generated_at"] = stamp

    out_dir.mkdir(parents=True, exist_ok=True)
    (out_dir / "summary.json").write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    (out_dir / "kol_pool_import_payload.json").write_text(
        json.dumps({"items": items, "source_type": "promo_plan_xlsx", "source_ref": xlsx_path.name}, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    (out_dir / "collaboration_records.json").write_text(json.dumps(collaboration_records, ensure_ascii=False, indent=2), encoding="utf-8")
    write_csv(out_dir / "kol_owner_mapping.csv", mapping_rows)
    write_csv(out_dir / "collaboration_records.csv", collaboration_records)

    result: dict[str, Any] | None = None
    if args.apply_kol_pool:
        if not args.staff_id:
            raise SystemExit("--apply-kol-pool requires --staff-id for audit attribution")
        result = apply_kol_pool(items, args.staff_id, xlsx_path.name)
        (out_dir / "kol_pool_apply_result.json").write_text(json.dumps(result, ensure_ascii=False, indent=2, default=str), encoding="utf-8")

    apply_summary = {}
    if result:
        apply_summary = {
            "imported": result.get("imported", 0),
            "skipped": result.get("skipped", 0),
            "items": len(result.get("items") or []),
        }
    print(json.dumps({"out_dir": str(out_dir), "summary": summary, "applied": bool(result), "apply_result": apply_summary}, ensure_ascii=False, indent=2, default=str))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
